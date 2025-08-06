from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
from django.conf import settings

class Command(BaseCommand):
    help = 'Create a sample Google Sheets template for lead import'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output',
            type=str,
            default='google_sheets_template.xlsx',
            help='Output filename for the template'
        )

    def handle(self, *args, **options):
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads Template"

        # Define headers
        required_headers = [
            'student_name',
            'parent_name', 
            'mobile_number',
            'email',
            'address',
            'block',
            'location_panchayat',
            'inquiry_source',
            'student_class'
        ]

        optional_headers = [
            'status',
            'remarks',
            'inquiry_date',
            'registration_date',
            'admission_test_date',
            'admission_offered_date',
            'admission_confirmed_date',
            'rejected_date',
            'follow_up_date'
        ]

        all_headers = required_headers + optional_headers

        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        required_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        optional_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Add headers
        for col, header in enumerate(all_headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_alignment
            
            # Color code required vs optional headers
            if header in required_headers:
                cell.fill = required_fill
            else:
                cell.fill = optional_fill

        # Add sample data
        sample_data = [
            ['John Doe', 'Jane Doe', '9876543210', 'john.doe@email.com', '123 Main St, City', 'Block A', 'Panchayat 1', 'Advertisement', 'Grade 5', 'Inquiry', 'Interested in admission', '2024-01-15', '', '', '', '', '', '2024-01-20'],
            ['Sarah Smith', 'Mike Smith', '9876543211', 'sarah.smith@email.com', '456 Oak Ave, Town', 'Block B', 'Panchayat 2', 'Walk-in', 'Grade 3', 'Registration', 'Completed registration', '2024-01-10', '2024-01-12', '', '', '', '', '2024-01-25'],
            ['Alex Johnson', 'Lisa Johnson', '9876543212', 'alex.johnson@email.com', '789 Pine Rd, Village', 'Block C', 'Panchayat 3', 'Online Form', 'Grade 7', 'Admission Test', 'Scheduled for test', '2024-01-05', '2024-01-08', '2024-01-18', '', '', '', '2024-01-22'],
        ]

        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = center_alignment

        # Add instructions
        instructions_row = len(sample_data) + 3
        ws.cell(row=instructions_row, column=1, value="INSTRUCTIONS:").font = Font(bold=True)
        ws.cell(row=instructions_row + 1, column=1, value="1. Red headers are REQUIRED fields")
        ws.cell(row=instructions_row + 2, column=1, value="2. Orange headers are OPTIONAL fields")
        ws.cell(row=instructions_row + 3, column=1, value="3. Date format: YYYY-MM-DD (e.g., 2024-01-15)")
        ws.cell(row=instructions_row + 4, column=1, value="4. inquiry_source options: Advertisement, Walk-in, Online Form, Referral")
        ws.cell(row=instructions_row + 5, column=1, value="5. student_class options: Play School, Nursery, LKG, UKG, Grade 1-12")
        ws.cell(row=instructions_row + 6, column=1, value="6. status options: Inquiry, Registration, Admission Test, Admission Offered, Admission Confirmed, Rejected")
        ws.cell(row=instructions_row + 7, column=1, value="7. block and location_panchayat should match your system's available options")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save the file
        output_path = os.path.join(settings.BASE_DIR, options['output'])
        wb.save(output_path)
        
        self.stdout.write(
            self.style.SUCCESS(f'Successfully created Google Sheets template: {output_path}')
        )
        self.stdout.write(
            self.style.WARNING('Upload this file to Google Sheets and share it with your application credentials.')
        ) 